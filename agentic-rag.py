import os
import json
import time
import asyncio
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from pathlib import Path

# Core dependencies
import psycopg2
from psycopg2.extras import RealDictCursor
import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer
import openai
from supabase import create_client, Client
from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

# File processing
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import PyPDF2
import docx
import openpyxl
import csv
from langchain.text_splitter import CharacterTextSplitter

from dotenv import load_dotenv
load_dotenv()


# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
@dataclass
class Config:
    # Database
    # postgres_url: str = os.getenv("POSTGRES_URL", "postgresql://user:password@localhost/dbname")
    postgres_url: str = os.getenv("POSTGRES_URL")

    # Supabase
    supabase_url: str = os.getenv("SUPABASE_URL")
    supabase_key: str = os.getenv("SUPABASE_KEY")
    
    # OpenAI
    # openai_api_key: str = os.getenv("OPENAI_API_KEY")
    
    # Mistral
    mistral_api_key: str = os.getenv("MISTRAL_API_KEY")
        
    # Google Drive
    # google_credentials_path: str = os.getenv("GOOGLE_CREDENTIALS_PATH", "credentials.json")
    # google_folder_id: str = os.getenv("GOOGLE_FOLDER_ID", "1ac8bnwRfKt5EuitYWiglHm848lrBUP9_")
    google_credentials_path: str = os.getenv("GOOGLE_CREDENTIALS_PATH")
    google_folder_id: str = os.getenv("GOOGLE_FOLDER_ID")
    
    # Other settings
    embedding_model: str = "text-embedding-3-small"
    chunk_size: int = 1000
    chunk_overlap: int = 200

config = Config()

# Pydantic models for API
class ChatRequest(BaseModel):
    chatInput: str
    sessionId: str

class ChatResponse(BaseModel):
    response: str
    sessionId: str

# Database setup and management
class DatabaseManager:
    def __init__(self, postgres_url: str):
        self.postgres_url = postgres_url
        
    def get_connection(self):
        return psycopg2.connect(self.postgres_url)
    
    def setup_tables(self):
        """Create all necessary database tables"""
        with self.get_connection() as conn:
            with conn.cursor() as cur:
                # Enable pgvector extension
                cur.execute("CREATE EXTENSION IF NOT EXISTS vector;")
                
                # Create documents table for vector storage
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS documents (
                        id BIGSERIAL PRIMARY KEY,
                        content TEXT,
                        metadata JSONB,
                        embedding VECTOR(1536)
                    );
                """)
                
                # Create document metadata table
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS document_metadata (
                        id TEXT PRIMARY KEY,
                        title TEXT,
                        url TEXT,
                        created_at TIMESTAMP DEFAULT NOW(),
                        schema TEXT
                    );
                """)
                
                # Create document rows table for tabular data
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS document_rows (
                        id SERIAL PRIMARY KEY,
                        dataset_id TEXT REFERENCES document_metadata(id),
                        row_data JSONB
                    );
                """)
                
                # Create chat memory table
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS chat_memory (
                        id SERIAL PRIMARY KEY,
                        session_id TEXT,
                        message_type TEXT,
                        content TEXT,
                        created_at TIMESTAMP DEFAULT NOW()
                    );
                """)
                
                # Create match function for vector similarity search
                cur.execute("""
                    CREATE OR REPLACE FUNCTION match_documents (
                        query_embedding VECTOR(1536),
                        match_count INT DEFAULT NULL,
                        filter JSONB DEFAULT '{}'
                    ) RETURNS TABLE (
                        id BIGINT,
                        content TEXT,
                        metadata JSONB,
                        similarity FLOAT
                    )
                    LANGUAGE plpgsql
                    AS $$
                    #variable_conflict use_column
                    BEGIN
                        RETURN QUERY
                        SELECT
                            id,
                            content,
                            metadata,
                            1 - (documents.embedding <=> query_embedding) AS similarity
                        FROM documents
                        WHERE metadata @> filter
                        ORDER BY documents.embedding <=> query_embedding
                        LIMIT match_count;
                    END;
                    $$;
                """)
                
                conn.commit()
                logger.info("Database tables created successfully")

# File processing utilities
class FileProcessor:
    def __init__(self):
        self.text_splitter = CharacterTextSplitter(
            chunk_size=config.chunk_size,
            chunk_overlap=config.chunk_overlap
        )
    
    def extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file"""
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting PDF text: {e}")
            return ""
    
    def extract_text_from_docx(self, file_content: bytes) -> str:
        """Extract text from DOCX file"""
        try:
            doc_file = io.BytesIO(file_content)
            doc = docx.Document(doc_file)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting DOCX text: {e}")
            return ""
    
    def extract_data_from_csv(self, file_content: bytes) -> List[Dict]:
        """Extract data from CSV file"""
        try:
            csv_file = io.StringIO(file_content.decode('utf-8'))
            reader = csv.DictReader(csv_file)
            return list(reader)
        except Exception as e:
            logger.error(f"Error extracting CSV data: {e}")
            return []
    
    def extract_data_from_excel(self, file_content: bytes) -> List[Dict]:
        """Extract data from Excel file"""
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file)
            all_data = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                data = []
                headers = [cell.value for cell in sheet[1]]
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    row_data['sheet_name'] = sheet_name
                    data.append(row_data)
                
                all_data.extend(data)
            
            return all_data
        except Exception as e:
            logger.error(f"Error extracting Excel data: {e}")
            return []

# Google Drive integration
class GoogleDriveMonitor:
    def __init__(self, credentials_path: str, folder_id: str):
        self.credentials_path = credentials_path
        self.folder_id = folder_id
        self.service = None
        self.file_processor = FileProcessor()
    
    def authenticate(self):
        """Authenticate with Google Drive API"""
        try:
            creds = Credentials.from_authorized_user_file(self.credentials_path)
            self.service = build('drive', 'v3', credentials=creds)
            logger.info("Google Drive authentication successful")
        except Exception as e:
            logger.error(f"Google Drive authentication failed: {e}")
            raise
    
    def get_files_in_folder(self) -> List[Dict]:
        """Get all files in the monitored folder"""
        try:
            results = self.service.files().list(
                q=f"'{self.folder_id}' in parents and trashed=false",
                fields="files(id,name,mimeType,modifiedTime,webViewLink)"
            ).execute()
            return results.get('files', [])
        except Exception as e:
            logger.error(f"Error getting files from folder: {e}")
            return []
    
    def download_file(self, file_id: str) -> bytes:
        """Download file content from Google Drive"""
        try:
            request = self.service.files().get_media(fileId=file_id)
            file_content = io.BytesIO()
            downloader = MediaIoBaseDownload(file_content, request)
            
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            
            return file_content.getvalue()
        except Exception as e:
            logger.error(f"Error downloading file {file_id}: {e}")
            return b""

# Vector store and embedding management
class VectorStore:
    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager
        openai.api_key = config.openai_api_key
    
    def get_embedding(self, text: str) -> List[float]:
        """Get embedding for text using OpenAI"""
        try:
            response = openai.Embedding.create(
                model=config.embedding_model,
                input=text
            )
            return response['data'][0]['embedding']
        except Exception as e:
            logger.error(f"Error getting embedding: {e}")
            return []
    
    def store_document_chunks(self, chunks: List[str], metadata: Dict, file_id: str):
        """Store document chunks with embeddings"""
        with self.db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # Delete old documents for this file
                cur.execute(
                    "DELETE FROM documents WHERE metadata->>'file_id' = %s",
                    (file_id,)
                )
                
                # Store new chunks
                for chunk in chunks:
                    embedding = self.get_embedding(chunk)
                    if embedding:
                        chunk_metadata = {**metadata, 'file_id': file_id}
                        cur.execute(
                            "INSERT INTO documents (content, metadata, embedding) VALUES (%s, %s, %s)",
                            (chunk, json.dumps(chunk_metadata), embedding)
                        )
                
                conn.commit()
                logger.info(f"Stored {len(chunks)} chunks for file {file_id}")
    
    def search_similar_documents(self, query: str, limit: int = 5) -> List[Dict]:
        """Search for similar documents"""
        query_embedding = self.get_embedding(query)
        if not query_embedding:
            return []
        
        with self.db_manager.get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT * FROM match_documents(%s, %s)",
                    (query_embedding, limit)
                )
                return [dict(row) for row in cur.fetchall()]

# Document processing and storage
class DocumentProcessor:
    def __init__(self, db_manager: DatabaseManager, vector_store: VectorStore):
        self.db_manager = db_manager
        self.vector_store = vector_store
        self.file_processor = FileProcessor()
    
    def process_file(self, file_info: Dict, file_content: bytes):
        """Process a file and store it in the database"""
        file_id = file_info['id']
        file_name = file_info['name']
        mime_type = file_info['mimeType']
        
        # Store file metadata
        self._store_file_metadata(file_info)
        
        # Process based on file type
        if mime_type == 'application/pdf':
            self._process_pdf(file_id, file_name, file_content)
        elif mime_type in ['application/vnd.google-apps.document', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
            self._process_document(file_id, file_name, file_content)
        elif mime_type in ['text/csv', 'application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            self._process_spreadsheet(file_id, file_name, file_content, mime_type)
    
    def _store_file_metadata(self, file_info: Dict):
        """Store file metadata in database"""
        with self.db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO document_metadata (id, title, url, created_at)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        title = EXCLUDED.title,
                        url = EXCLUDED.url
                    """,
                    (file_info['id'], file_info['name'], file_info.get('webViewLink', ''), datetime.now())
                )
                conn.commit()
    
    def _process_pdf(self, file_id: str, file_name: str, file_content: bytes):
        """Process PDF file"""
        text = self.file_processor.extract_text_from_pdf(file_content)
        if text:
            chunks = self.file_processor.text_splitter.split_text(text)
            metadata = {'file_title': file_name, 'file_type': 'pdf'}
            self.vector_store.store_document_chunks(chunks, metadata, file_id)
    
    def _process_document(self, file_id: str, file_name: str, file_content: bytes):
        """Process document file (DOCX, etc.)"""
        text = self.file_processor.extract_text_from_docx(file_content)
        if text:
            chunks = self.file_processor.text_splitter.split_text(text)
            metadata = {'file_title': file_name, 'file_type': 'document'}
            self.vector_store.store_document_chunks(chunks, metadata, file_id)
    
    def _process_spreadsheet(self, file_id: str, file_name: str, file_content: bytes, mime_type: str):
        """Process spreadsheet file (CSV, Excel)"""
        if 'csv' in mime_type:
            data = self.file_processor.extract_data_from_csv(file_content)
        else:
            data = self.file_processor.extract_data_from_excel(file_content)
        
        if data:
            # Store individual rows
            self._store_tabular_data(file_id, data)
            
            # Create text representation for vector search
            text_data = self._convert_tabular_to_text(data)
            chunks = self.file_processor.text_splitter.split_text(text_data)
            metadata = {'file_title': file_name, 'file_type': 'spreadsheet'}
            self.vector_store.store_document_chunks(chunks, metadata, file_id)
            
            # Store schema
            if data:
                schema = list(data[0].keys())
                self._update_file_schema(file_id, schema)
    
    def _store_tabular_data(self, file_id: str, data: List[Dict]):
        """Store tabular data in document_rows table"""
        with self.db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # Delete old rows
                cur.execute("DELETE FROM document_rows WHERE dataset_id = %s", (file_id,))
                
                # Insert new rows
                for row in data:
                    cur.execute(
                        "INSERT INTO document_rows (dataset_id, row_data) VALUES (%s, %s)",
                        (file_id, json.dumps(row))
                    )
                
                conn.commit()
    
    def _convert_tabular_to_text(self, data: List[Dict]) -> str:
        """Convert tabular data to text for vector search"""
        text_parts = []
        for row in data:
            row_text = ", ".join([f"{k}: {v}" for k, v in row.items() if v is not None])
            text_parts.append(row_text)
        return "\n".join(text_parts)
    
    def _update_file_schema(self, file_id: str, schema: List[str]):
        """Update file schema in metadata"""
        with self.db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE document_metadata SET schema = %s WHERE id = %s",
                    (json.dumps(schema), file_id)
                )
                conn.commit()

# Chat and RAG functionality
class RAGAgent:
    def __init__(self, db_manager: DatabaseManager, vector_store: VectorStore):
        self.db_manager = db_manager
        self.vector_store = vector_store
        openai.api_key = config.openai_api_key
    
    def get_chat_response(self, query: str, session_id: str) -> str:
        """Get response for user query using RAG"""
        # First, try RAG search
        similar_docs = self.vector_store.search_similar_documents(query, limit=5)
        
        # Get chat history
        chat_history = self._get_chat_history(session_id)
        
        # Prepare context
        context = self._prepare_context(similar_docs)
        
        # Generate response
        response = self._generate_response(query, context, chat_history)
        
        # Store conversation
        self._store_chat_message(session_id, "human", query)
        self._store_chat_message(session_id, "assistant", response)
        
        return response
    
    def _get_chat_history(self, session_id: str) -> List[Dict]:
        """Get chat history for session"""
        with self.db_manager.get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT message_type, content FROM chat_memory 
                    WHERE session_id = %s 
                    ORDER BY created_at DESC 
                    LIMIT 10
                    """,
                    (session_id,)
                )
                return [dict(row) for row in reversed(cur.fetchall())]
    
    def _prepare_context(self, similar_docs: List[Dict]) -> str:
        """Prepare context from similar documents"""
        if not similar_docs:
            return "No relevant documents found."
        
        context_parts = []
        for doc in similar_docs:
            context_parts.append(f"Document: {doc['content'][:500]}...")
        
        return "\n\n".join(context_parts)
    
    def _generate_response(self, query: str, context: str, chat_history: List[Dict]) -> str:
        """Generate response using OpenAI"""
        # Prepare conversation history
        messages = [
            {
                "role": "system",
                "content": """You are a personal assistant who helps answer questions from a corpus of documents. 
                The documents are either text based (Txt, docs, extracted PDFs, etc.) or tabular data (CSVs or Excel documents).
                
                Always start by performing RAG unless the question requires a SQL query for tabular data. 
                If RAG doesn't help, then look at the documents that are available to you, find a few that you think would contain the answer, and then analyze those.
                
                Always tell the user if you didn't find the answer. Don't make something up just to please them."""
            }
        ]
        
        # Add chat history
        for msg in chat_history[-5:]:  # Last 5 messages
            role = "user" if msg['message_type'] == "human" else "assistant"
            messages.append({"role": role, "content": msg['content']})
        
        # Add current query with context
        user_message = f"Context from documents:\n{context}\n\nUser question: {query}"
        messages.append({"role": "user", "content": user_message})
        
        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages,
                max_tokens=1000,
                temperature=0.7
            )
            return response.choices[0].message.content
        except Exception as e:
            logger.error(f"Error generating response: {e}")
            return "Sorry, I encountered an error while processing your request."
    
    def _store_chat_message(self, session_id: str, message_type: str, content: str):
        """Store chat message in database"""
        with self.db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO chat_memory (session_id, message_type, content) VALUES (%s, %s, %s)",
                    (session_id, message_type, content)
                )
                conn.commit()

# File monitoring service
class FileMonitorService:
    def __init__(self, google_drive_monitor: GoogleDriveMonitor, document_processor: DocumentProcessor):
        self.google_drive_monitor = google_drive_monitor
        self.document_processor = document_processor
        self.processed_files = set()
        self.running = False
    
    async def start_monitoring(self):
        """Start monitoring Google Drive folder"""
        self.running = True
        logger.info("Starting file monitoring service")
        
        while self.running:
            try:
                files = self.google_drive_monitor.get_files_in_folder()
                
                for file_info in files:
                    file_id = file_info['id']
                    modified_time = file_info['modifiedTime']
                    
                    # Create unique identifier for file + modification time
                    file_signature = f"{file_id}_{modified_time}"
                    
                    if file_signature not in self.processed_files:
                        logger.info(f"Processing new/updated file: {file_info['name']}")
                        
                        # Download and process file
                        file_content = self.google_drive_monitor.download_file(file_id)
                        if file_content:
                            self.document_processor.process_file(file_info, file_content)
                            self.processed_files.add(file_signature)
                
                # Wait before next check
                await asyncio.sleep(60)  # Check every minute
                
            except Exception as e:
                logger.error(f"Error in file monitoring: {e}")
                await asyncio.sleep(60)
    
    def stop_monitoring(self):
        """Stop monitoring"""
        self.running = False
        logger.info("File monitoring service stopped")

# FastAPI application
app = FastAPI(title="RAG System API", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure appropriately for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global instances
db_manager = None
vector_store = None
rag_agent = None
file_monitor_service = None

@app.on_event("startup")
async def startup_event():
    """Initialize services on startup"""
    global db_manager, vector_store, rag_agent, file_monitor_service
    
    try:
        # Initialize database
        db_manager = DatabaseManager(config.postgres_url)
        db_manager.setup_tables()
        
        # Initialize vector store
        vector_store = VectorStore(db_manager)
        
        # Initialize RAG agent
        rag_agent = RAGAgent(db_manager, vector_store)
        
        # Initialize file monitoring
        google_drive_monitor = GoogleDriveMonitor(
            config.google_credentials_path, 
            config.google_folder_id
        )
        google_drive_monitor.authenticate()
        
        document_processor = DocumentProcessor(db_manager, vector_store)
        file_monitor_service = FileMonitorService(google_drive_monitor, document_processor)
        
        # Start file monitoring in background
        asyncio.create_task(file_monitor_service.start_monitoring())
        
        logger.info("RAG System initialized successfully")
        
    except Exception as e:
        logger.error(f"Failed to initialize RAG System: {e}")
        raise

@app.post("/chat", response_model=ChatResponse)
async def chat_endpoint(request: ChatRequest):
    """Chat endpoint for RAG queries"""
    try:
        response = rag_agent.get_chat_response(request.chatInput, request.sessionId)
        return ChatResponse(response=response, sessionId=request.sessionId)
    except Exception as e:
        logger.error(f"Error in chat endpoint: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/documents")
async def list_documents():
    """List all documents in the system"""
    try:
        with db_manager.get_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT * FROM document_metadata ORDER BY created_at DESC")
                return [dict(row) for row in cur.fetchall()]
    except Exception as e:
        logger.error(f"Error listing documents: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

# Main execution
if __name__ == "__main__":
    # Run the FastAPI server
    uvicorn.run(
        "agentic-rag:app",  
        host="0.0.0.0",
        port=8000,
        reload=True
    )